# gui/moderator_window.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import datetime, time
from config import Config
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class ModeratorWindow:
    def __init__(self, root, db, user, logout_callback):
        self.root = root
        self.db = db
        self.user = user
        self.logout_callback = logout_callback
        
        self.root.title(f"{Config.WINDOW_TITLE} - Kiểm Duyệt")
        self.root.geometry(f"{Config.WINDOW_WIDTH}x{Config.WINDOW_HEIGHT}")
        
        self.center_window()
        self.create_widgets()
        self.refresh_classes()
        self.load_registration_period()

    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    # =================== TẠO WIDGET ===================
    def create_widgets(self):
        # Header
        header = tk.Frame(self.root, bg='#667eea', height=80)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        tk.Label(
            header,
            text=f"🔍 Chào mừng, {self.user['full_name']}",
            font=('Arial', 16, 'bold'),
            bg='#667eea',
            fg='white'
        ).pack(side=tk.LEFT, padx=20, pady=20)
        
        tk.Button(
            header,
            text="🚪 Đăng xuất",
            font=('Arial', 11),
            bg='white',
            fg='#667eea',
            cursor='hand2',
            command=self.logout
        ).pack(side=tk.RIGHT, padx=20)
        
        # Notebook cho tab
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Các tab
        self.create_classes_tab()
        self.create_registration_period_tab()
        
        # Tab quản lý tài khoản
        self.create_account_management_tab('student')
        self.create_account_management_tab('teacher')

    # =================== TAB LỚP HỌC ===================
    def create_classes_tab(self):
        tab = tk.Frame(self.notebook, bg='white')
        self.notebook.add(tab, text="📋 Duyệt Lớp Học")
        
        filter_frame = tk.Frame(tab, bg='white')
        filter_frame.pack(fill=tk.X, padx=10, pady=10)
        
        tk.Label(filter_frame, text="Học kỳ:", bg='white').pack(side=tk.LEFT, padx=5)
        self.semester_combo = ttk.Combobox(filter_frame, values=['Tất cả', '1', '2', '3'], width=10)
        self.semester_combo.set('Tất cả')
        self.semester_combo.pack(side=tk.LEFT, padx=5)
        
        tk.Label(filter_frame, text="Năm học:", bg='white').pack(side=tk.LEFT, padx=5)
        self.year_entry = tk.Entry(filter_frame, width=15)
        self.year_entry.pack(side=tk.LEFT, padx=5)
        
        tk.Button(filter_frame, text="🔍 Lọc", bg='#667eea', fg='white',
                  command=self.filter_classes).pack(side=tk.LEFT, padx=5)
        tk.Button(filter_frame, text="🔄 Làm mới", bg='#e0e0e0',
                  command=self.refresh_classes).pack(side=tk.LEFT, padx=5)
        
        tree_frame = tk.Frame(tab, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        y_scroll = ttk.Scrollbar(tree_frame)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        x_scroll = ttk.Scrollbar(tree_frame, orient='horizontal')
        x_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.classes_tree = ttk.Treeview(
            tree_frame,
            columns=('id','code','name','teacher','credits','students','semester','year','status','created'),
            show='headings',
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set
        )
        y_scroll.config(command=self.classes_tree.yview)
        x_scroll.config(command=self.classes_tree.xview)
        
        headers = {
            'id': ('ID', 50),
            'code': ('Mã lớp', 100),
            'name': ('Tên lớp', 200),
            'teacher': ('Giảng viên', 150),
            'credits': ('Tín chỉ', 70),
            'students': ('SV', 70),
            'semester': ('Kỳ', 50),
            'year': ('Năm học', 100),
            'status': ('Trạng thái', 120),
            'created': ('Ngày tạo', 150)
        }
        for col, (text, width) in headers.items():
            self.classes_tree.heading(col, text=text)
            self.classes_tree.column(col, width=width)
        self.classes_tree.pack(fill=tk.BOTH, expand=True)
        
        action_frame = tk.Frame(tab, bg='white')
        action_frame.pack(fill=tk.X, pady=10)
        
        tk.Button(action_frame, text="✓ Duyệt lớp", bg='#28a745', fg='white',
                  command=self.approve_selected, width=20).pack(side=tk.LEFT, padx=10)
        tk.Button(action_frame, text="✗ Từ chối lớp", bg='#dc3545', fg='white',
                  command=self.reject_selected, width=20).pack(side=tk.LEFT, padx=10)
        
        self.classes_tree.bind('<Button-3>', self.show_context_menu)

    # =================== TAB QUẢN LÝ THỜI GIAN ===================
    def create_registration_period_tab(self):
        tab = tk.Frame(self.notebook, bg='white')
        self.notebook.add(tab, text="⏰ Quản lý thời gian đăng ký")
        
        frame = tk.Frame(tab, bg='white')
        frame.pack(padx=20, pady=20, anchor='w')
        
        tk.Label(frame, text="Học kỳ:", bg='white', font=('Arial', 11, 'bold')).grid(row=0, column=0, padx=5, pady=10, sticky='w')
        self.period_semester = ttk.Combobox(frame, values=['1', '2', '3'], width=10, state='readonly')
        self.period_semester.grid(row=0, column=1, padx=5, pady=10, sticky='w')
        self.period_semester.current(0)
        
        tk.Label(frame, text="Năm học:", bg='white', font=('Arial', 11, 'bold')).grid(row=0, column=2, padx=15, pady=10, sticky='w')
        self.period_year = tk.Entry(frame, width=12)
        self.period_year.grid(row=0, column=3, padx=5, pady=10, sticky='w')
        self.period_year.insert(0, "2024-2025")
        
        tk.Label(frame, text="Bắt đầu:", bg='white', font=('Arial', 11)).grid(row=1, column=0, padx=5, pady=5, sticky='w')
        self.start_date = DateEntry(frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        self.start_date.grid(row=1, column=1, padx=5)
        self.start_hour = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(24)], width=3)
        self.start_hour.grid(row=1, column=2, padx=2)
        self.start_hour.set("08")
        self.start_minute = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(0,60,5)], width=3)
        self.start_minute.grid(row=1, column=3, padx=2)
        self.start_minute.set("00")
        
        tk.Label(frame, text="Kết thúc:", bg='white', font=('Arial', 11)).grid(row=2, column=0, padx=5, pady=5, sticky='w')
        self.end_date = DateEntry(frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        self.end_date.grid(row=2, column=1, padx=5)
        self.end_hour = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(24)], width=3)
        self.end_hour.grid(row=2, column=2, padx=2)
        self.end_hour.set("23")
        self.end_minute = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(0,60,5)], width=3)
        self.end_minute.grid(row=2, column=3, padx=2)
        self.end_minute.set("59")
        
        tk.Button(frame, text="💾 Lưu thời gian đăng ký", bg="#28a745", fg="white",
                  font=('Arial', 12, 'bold'), command=self.save_registration_period).grid(row=3, column=0, columnspan=4, pady=20)

    # =================== TAB QUẢN LÝ TÀI KHOẢN CHUNG ===================
    def create_account_management_tab(self, account_type='student'):
        if account_type=='student':
            title = "👥 Cấp tài khoản sinh viên"
            columns = ['code','name','gender','dob','major','year']
            headers = {'code':'MSSV','name':'Họ tên','gender':'Giới tính','dob':'Ngày sinh','major':'Ngành học','year':'Khóa học'}
            email_domain = 'student.edu.vn'
        else:
            title = "👩‍🏫 Cấp tài khoản giảng viên"
            columns = ['code','name','gender','dob','faculty']
            headers = {'code':'Mã GV','name':'Họ tên','gender':'Giới tính','dob':'Ngày sinh','faculty':'Thuộc khoa'}
            email_domain = 'faculty.edu.vn'

        tab = tk.Frame(self.notebook, bg='white')
        self.notebook.add(tab, text=title)

        info_frame = tk.LabelFrame(tab, text="📖 Hướng dẫn", bg='white', font=('Arial', 11, 'bold'))
        info_frame.pack(fill=tk.X, padx=20, pady=10)
        info_text = f"""
• Chuẩn bị file Excel có các cột: {', '.join(headers.values())}
• Giới tính: male, female, other
• Ngày sinh: định dạng YYYY-MM-DD
• Hệ thống sẽ tự tạo: Username = {headers['code']}, Password = {headers['code']}, Email = {headers['code']}@{email_domain}
"""
        tk.Label(info_frame, text=info_text, bg='white', justify='left', font=('Arial', 10)).pack(padx=10, pady=10)

        input_frame = tk.Frame(tab, bg='white'); input_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Button(input_frame, text="📂 Chọn file Excel", bg='#007bff', fg='white', font=('Arial', 12, 'bold'),
                  command=lambda:self.select_account_file(account_type), width=20).pack(side=tk.LEFT, padx=10)
        tk.Button(input_frame, text="➕ Thêm thủ công", bg='#17a2b8', fg='white', font=('Arial', 12, 'bold'),
                  command=lambda:self.add_account_manual(account_type), width=25).pack(side=tk.LEFT, padx=10)

        tree_frame = tk.Frame(tab, bg='white'); tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        tk.Label(tree_frame, text=f"Danh sách {account_type} sẽ tạo:", bg='white', font=('Arial', 11, 'bold')).pack(anchor='w', pady=5)
        treeview = ttk.Treeview(tree_frame, columns=columns, show='headings', height=10)
        for col in columns:
            treeview.heading(col, text=headers[col])
            treeview.column(col, width=150)
        y_scroll = ttk.Scrollbar(tree_frame, orient='vertical', command=treeview.yview)
        treeview.configure(yscrollcommand=y_scroll.set)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        treeview.pack(fill=tk.BOTH, expand=True)

        action_frame = tk.Frame(tab, bg='white'); action_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Button(action_frame, text="✓ Tạo tài khoản", bg='#28a745', fg='white', font=('Arial', 12, 'bold'),
                  command=lambda:self.create_accounts(account_type), width=20).pack(side=tk.LEFT, padx=10)
        tk.Button(action_frame, text="📥 Xuất danh sách vừa tạo", bg='#ffc107', fg='black', font=('Arial', 11, 'bold'),
                  command=lambda:self.export_accounts_excel(account_type), width=25).pack(side=tk.LEFT, padx=10)
        tk.Button(action_frame, text="🗑️ Xóa danh sách", bg='#dc3545', fg='white', font=('Arial', 12, 'bold'),
                  command=lambda:self.clear_account_list(account_type), width=20).pack(side=tk.LEFT, padx=10)
        
        # Thêm frame mới cho xuất database
        export_frame = tk.Frame(tab, bg='white'); export_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Label(export_frame, text="📊 Xuất dữ liệu từ Database:", bg='white', font=('Arial', 11, 'bold')).pack(side=tk.LEFT, padx=10)
        tk.Button(export_frame, text="📤 Xuất tất cả từ Database", bg='#17a2b8', fg='white', font=('Arial', 12, 'bold'),
                  command=lambda:self.export_all_from_database(account_type), width=30).pack(side=tk.LEFT, padx=10)

        setattr(self, f"pending_{account_type}s", [])
        setattr(self, f"created_{account_type}_accounts", [])
        setattr(self, f"{account_type}_tree", treeview)

    # =================== CHUNG: SELECT FILE & ADD MANUAL ===================
    def select_account_file(self, account_type):
        file_path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel files","*.xlsx *.xls")])
        if not file_path: return

        pending_list = getattr(self, f"pending_{account_type}s")
        treeview = getattr(self, f"{account_type}_tree")

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            if account_type=='student':
                keys = ['student_code','full_name','gender','date_of_birth','major','enrollment_year']
            else:
                keys = ['teacher_code','full_name','gender','date_of_birth','faculty']

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                account = {k: (row[i] if i < len(row) and row[i] is not None else '') for i, k in enumerate(keys)}
                pending_list.append(account)
                treeview.insert('', tk.END, values=tuple(account.values()))
            messagebox.showinfo("Thành công", f"Đã tải {len(pending_list)} {account_type}!")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file: {e}")

    def add_account_manual(self, account_type):
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Thêm {account_type}")
        dialog.geometry("400x350")
        dialog.transient(self.root)
        dialog.grab_set()

        fields = {}
        if account_type=='student':
            labels = [("MSSV:",'code'), ("Họ tên:",'full_name'), ("Giới tính:",'gender'),
                      ("Ngày sinh (YYYY-MM-DD):",'date_of_birth'), ("Ngành học:",'major'), ("Khóa học:",'year')]
        else:
            labels = [("Mã GV:",'code'), ("Họ tên:",'full_name'), ("Giới tính:",'gender'),
                      ("Ngày sinh (YYYY-MM-DD):",'date_of_birth'), ("Thuộc khoa:",'faculty')]

        for i, (label_text, key) in enumerate(labels):
            tk.Label(dialog, text=label_text).grid(row=i, column=0, padx=10, pady=5, sticky='w')
            if key=='gender':
                fields[key] = ttk.Combobox(dialog, values=['male','female','other'], width=25)
                fields[key].current(0)
            else:
                fields[key] = tk.Entry(dialog, width=28)
            fields[key].grid(row=i, column=1, padx=10, pady=5)

        def save_account():
            account = {key: fields[key].get().strip() for _, key in labels}
            if not account[list(account.keys())[0]] or not account['full_name']:
                messagebox.showerror("Lỗi", "ID và Họ tên không được để trống")
                return
            pending_list = getattr(self, f"pending_{account_type}s")
            treeview = getattr(self, f"{account_type}_tree")
            pending_list.append(account)
            treeview.insert('', tk.END, values=tuple(account.values()))
            dialog.destroy()

        tk.Button(dialog, text="✓ Thêm", bg='#28a745', fg='white', command=save_account).grid(
            row=len(labels), column=0, columnspan=2, pady=20
        )

    # =================== CREATE & EXPORT ACCOUNTS ===================
    def create_accounts(self, account_type):
        pending_list = getattr(self, f"pending_{account_type}s")
        if not pending_list:
            messagebox.showwarning("Cảnh báo",f"Chưa có {account_type} nào!")
            return
        if not messagebox.askyesno("Xác nhận", f"Tạo {len(pending_list)} tài khoản {account_type}?"):
            return
        func = {'student': self.db.create_students_bulk, 'teacher': self.db.create_teachers_bulk}[account_type]
        success_count, created = func(pending_list)
        setattr(self, f"created_{account_type}_accounts", created)
        messagebox.showinfo("Thành công", f"Đã tạo {success_count} tài khoản {account_type}!")

    def export_accounts_excel(self, account_type):
        created_accounts = getattr(self, f"created_{account_type}_accounts")
        if not created_accounts:
            messagebox.showwarning("Cảnh báo", f"Chưa có {account_type} nào để xuất!")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")])
        if not file_path: return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Accounts"
        keys = list(created_accounts[0].keys())
        for col, key in enumerate(keys, start=1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="00C0C0C0")
            cell.alignment = Alignment(horizontal='center')
        for r, account in enumerate(created_accounts, start=2):
            for c, key in enumerate(keys, start=1):
                ws.cell(row=r, column=c, value=account[key])
        wb.save(file_path)
        messagebox.showinfo("Thành công", f"Xuất Excel thành công: {file_path}")

    def clear_account_list(self, account_type):
        pending_list = getattr(self, f"pending_{account_type}s")
        treeview = getattr(self, f"{account_type}_tree")
        if not pending_list: return
        if not messagebox.askyesno("Xác nhận", f"Xóa danh sách {account_type}?"): return
        pending_list.clear()
        for item in treeview.get_children():
            treeview.delete(item)
    
    def export_all_from_database(self, account_type):
        """Xuất tất cả sinh viên/giảng viên từ database ra Excel"""
        try:
            # Lấy dữ liệu từ database
            if account_type == 'student':
                data = self.db.get_all_students()
                title = "Danh Sách Sinh Viên"
                columns = [
                    ('STT', 'stt'),
                    ('Mã sinh viên', 'student_code'),
                    ('Họ và tên', 'full_name'),
                    ('Giới tính', 'gender'),
                    ('Ngày sinh', 'date_of_birth'),
                    ('Email', 'email'),
                    ('Ngành học', 'major'),
                    ('Khóa học', 'enrollment_year'),
                    ('Có ảnh khuôn mặt', 'has_face')
                ]
            else:  # teacher
                data = self.db.get_all_teachers()
                title = "Danh Sách Giảng Viên"
                columns = [
                    ('STT', 'stt'),
                    ('Mã giảng viên', 'teacher_code'),
                    ('Họ và tên', 'full_name'),
                    ('Giới tính', 'gender'),
                    ('Ngày sinh', 'date_of_birth'),
                    ('Email', 'email'),
                    ('Khoa/Bộ môn', 'department')
                ]
            
            if not data:
                messagebox.showwarning("Cảnh báo", f"Không có {account_type} nào trong database!")
                return
            
            # Chọn nơi lưu file
            default_filename = f"DanhSach_{account_type.capitalize()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=default_filename
            )
            
            if not file_path:
                return
            
            # Tạo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title
            
            # Style cho header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Viết header
            for col_idx, (header_text, _) in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Viết dữ liệu
            for row_idx, item in enumerate(data, start=2):
                for col_idx, (_, key) in enumerate(columns, start=1):
                    if key == 'stt':
                        value = row_idx - 1
                    elif key == 'has_face':
                        value = "Có" if item.get('face_encoding_path') else "Chưa"
                    elif key == 'date_of_birth':
                        dob = item.get(key)
                        value = dob.strftime('%d/%m/%Y') if dob else ''
                    elif key == 'gender':
                        gender_map = {'male': 'Nam', 'female': 'Nữ', 'other': 'Khác'}
                        value = gender_map.get(item.get(key, ''), item.get(key, ''))
                    else:
                        value = item.get(key, '')
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Tự động điều chỉnh độ rộng cột
            for col_idx, (header_text, _) in enumerate(columns, start=1):
                max_length = len(header_text)
                for row_idx in range(2, len(data) + 2):
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value)
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
            
            # Thêm thông tin tổng kết
            summary_row = len(data) + 3
            ws.cell(row=summary_row, column=1, value=f"Tổng số: {len(data)} {account_type}")
            ws.cell(row=summary_row, column=1).font = Font(bold=True, size=11)
            
            ws.cell(row=summary_row + 1, column=1, value=f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            ws.cell(row=summary_row + 1, column=1).font = Font(italic=True, size=10)
            
            # Lưu file
            wb.save(file_path)
            
            messagebox.showinfo(
                "Thành công",
                f"Đã xuất {len(data)} {account_type} ra file:\n{file_path}"
            )
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất Excel:\n{str(e)}")

    # =================== LOGOUT ===================
    def logout(self):
        if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn đăng xuất?"):
            self.logout_callback()

    # =================== LOAD DANH SÁCH LỚP ===================
    def refresh_classes(self):
        for item in self.classes_tree.get_children():
            self.classes_tree.delete(item)

        try:
            classes = self.db.get_classes_for_approval()
            for c in classes:
                self.classes_tree.insert(
                    '', tk.END,
                    values=(
                        c['class_id'],
                        c['class_code'],
                        c['class_name'],
                        c['teacher_name'],
                        c['credits'],
                        c['max_students'],
                        c['semester'],
                        c['academic_year'],
                        c['status'],
                        c['created_at']
                    )
                )
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không tải được lớp học:\n{e}")


    def filter_classes(self):
        semester = self.semester_combo.get()
        year = self.year_entry.get().strip()

        semester = None if semester == "Tất cả" else int(semester)
        year = None if not year else year

        for item in self.classes_tree.get_children():
            self.classes_tree.delete(item)

        try:
            classes = self.db.get_classes_for_approval(semester, year)
            for c in classes:
                self.classes_tree.insert(
                    '', tk.END,
                    values=(
                        c['class_id'],
                        c['class_code'],
                        c['class_name'],
                        c['teacher_name'],
                        c['credits'],
                        c['max_students'],
                        c['semester'],
                        c['academic_year'],
                        c['status'],
                        c['created_at']
                    )
                )
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi lọc lớp:\n{e}")


    # =================== DUYỆT & TỪ CHỐI ===================
    def approve_selected(self):
        selected = self.classes_tree.selection()
        if not selected:
            messagebox.showwarning("Chọn lớp", "Vui lòng chọn lớp cần duyệt")
            return

        class_id = self.classes_tree.item(selected[0])['values'][0]

        if self.db.approve_class(class_id):
            messagebox.showinfo("Thành công", "Đã duyệt lớp")
            self.refresh_classes()
        else:
            messagebox.showerror("Lỗi", "Không thể duyệt lớp")


    def reject_selected(self):
        selected = self.classes_tree.selection()
        if not selected:
            messagebox.showwarning("Chọn lớp", "Vui lòng chọn lớp cần từ chối")
            return

        class_id = self.classes_tree.item(selected[0])['values'][0]

        if self.db.reject_class(class_id):
            messagebox.showinfo("Thành công", "Đã từ chối lớp")
            self.refresh_classes()
        else:
            messagebox.showerror("Lỗi", "Không thể từ chối lớp")


    # =================== MENU CHUỘT PHẢI ===================
    def show_context_menu(self, event):
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="✓ Duyệt", command=self.approve_selected)
        menu.add_command(label="✗ Từ chối", command=self.reject_selected)
        menu.tk_popup(event.x_root, event.y_root)


    # =================== QUẢN LÝ THỜI GIAN ĐĂNG KÝ ===================
    def save_registration_period(self):
        try:
            start_dt = datetime.combine(
                self.start_date.get_date(),
                time(int(self.start_hour.get()), int(self.start_minute.get()))
            )
            end_dt = datetime.combine(
                self.end_date.get_date(),
                time(int(self.end_hour.get()), int(self.end_minute.get()))
            )

            semester = int(self.period_semester.get())
            year = self.period_year.get().strip()

            if start_dt >= end_dt:
                messagebox.showerror("Lỗi", "Thời gian bắt đầu phải trước kết thúc")
                return

            if self.db.save_registration_period(start_dt, end_dt, semester, year):
                messagebox.showinfo("Thành công", "Đã lưu thời gian đăng ký")
            else:
                messagebox.showerror("Lỗi", "Không thể lưu thời gian")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Dữ liệu không hợp lệ:\n{e}")


    def load_registration_period(self):
        data = self.db.get_latest_registration_window()
        if data:
            self.period_semester.set(str(data['semester']))
            self.period_year.delete(0, tk.END)
            self.period_year.insert(0, data['academic_year'])

            self.start_date.set_date(data['start_datetime'])
            self.end_date.set_date(data['end_datetime'])

            self.start_hour.set(f"{data['start_datetime'].hour:02d}")
            self.start_minute.set(f"{data['start_datetime'].minute:02d}")
            self.end_hour.set(f"{data['end_datetime'].hour:02d}")
            self.end_minute.set(f"{data['end_datetime'].minute:02d}")
