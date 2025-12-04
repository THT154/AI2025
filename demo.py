import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Dữ liệu ví dụ
teachers = [
    {'id': 'GV001', 'name': 'Nguyễn Văn A', 'gender': 'male', 'dob': '1980-05-10', 'department': 'Toán'},
    {'id': 'GV002', 'name': 'Trần Thị B', 'gender': 'female', 'dob': '1985-08-20', 'department': 'Lý'},
    {'id': 'GV003', 'name': 'Lê Văn C', 'gender': 'other', 'dob': '1990-12-15', 'department': 'Hóa'},
]

# Tạo file Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GiangVienAccounts"

# Header
headers = ['Mã GV', 'Họ tên', 'Giới tính', 'Ngày sinh', 'Bộ môn', 'Username', 'Password', 'Email']
ws.append(headers)

# Style header
fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Thêm dữ liệu giảng viên
for t in teachers:
    username = t['id']
    password = t['id']
    email = f"{t['id']}@faculty.edu.vn"
    ws.append([t['id'], t['name'], t['gender'], t['dob'], t['department'], username, password, email])

# Tự động chỉnh độ rộng cột
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2

# Lưu file
file_path = "GiangVienAccounts.xlsx"
wb.save(file_path)
print(f"Đã tạo file Excel: {file_path}")
