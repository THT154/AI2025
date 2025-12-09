# views/moderator_window_new.py - Cấu trúc mới theo yêu cầu
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import datetime, time
from config import Config
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class ModeratorWindow:
    def __init__(self, root, db, user, logout_callback):
        self.root = root
        self.db = db
        self.user = user
        self.logout_callback = logout_callback
        
        self.root.title(f"{Config.WINDOW_TITLE} - Kiểm Duyệt")
        self.root.geometry(f"{Config.WINDOW_WIDTH}x{Config.WINDOW_HEIGHT}")
        self.root.minsize(1000, 600)  # Kích thước tối thiểu
        self.root.resizable(True, True)  # Cho phép resize
        
        # Data storage
        self.pending_students = []
        self.pending_teachers = []
        self.created_student_accounts = []
        self.created_teacher_accounts = []
        
        self.center_window()
        self.create_widgets()
        self.refresh_classes()
        self.load_registration_period()
        
        # Xử lý đóng cửa sổ
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def create_widgets(self):
        # Header
        header = tk.Frame(self.root, bg='#667eea', height=80)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        tk.Label(
            header,
            text=f"🔍 Kiểm duyệt viên: {self.user['full_name']}",
            font=('Arial', 16, 'bold'),
            bg='#667eea',
            fg='white'
        ).pack(side=tk.LEFT, padx=20, pady=20)
        
        tk.Button(
            header,
            text="🚪 Đăng xuất",
            font=('Arial', 11),
            bg='white',
            fg='#667eea',
            cursor='hand2',
            command=self.logout
        ).pack(side=tk.RIGHT, padx=20)
        
        # Main Notebook
        self.main_notebook = ttk.Notebook(self.root)
        self.main_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Tab 1: Duyệt lớp học
        self.create_classes_tab()
        
        # Tab 2: Cấp tài khoản (có sub-tabs)
        self.create_account_creation_tab()
        
        # Tab 3: Quản lý tài khoản (có sub-tabs)
        self.create_account_management_tab()
        
        # Tab 4: Quản lý thời gian đăng ký
        self.create_registration_period_tab()
        
        # Tab 5: Train Model AI (MỚI)
        self.create_train_model_tab()
        
        # Tab 6: Yêu cầu hỗ trợ (MỚI)
        self.create_support_requests_tab()

    # =================== TAB 1: DUYỆT LỚP HỌC ===================
    def create_classes_tab(self):
        tab = tk.Frame(self.main_notebook, bg='white')
        self.main_notebook.add(tab, text="📋 Duyệt Lớp Học")
        
        # Filter frame
        filter_frame = tk.Frame(tab, bg='white')
        filter_frame.pack(fill=tk.X, padx=10, pady=10)
        
        tk.Label(filter_frame, text="Học kỳ:", bg='white').pack(side=tk.LEFT, padx=5)
        self.semester_combo = ttk.Combobox(filter_frame, values=['Tất cả', '1', '2', '3'], width=10)
        self.semester_combo.set('Tất cả')
        self.semester_combo.pack(side=tk.LEFT, padx=5)
        
        tk.Label(filter_frame, text="Năm học:", bg='white').pack(side=tk.LEFT, padx=5)
        self.year_entry = tk.Entry(filter_frame, width=15)
        self.year_entry.pack(side=tk.LEFT, padx=5)
        
        tk.Button(filter_frame, text="🔍 Lọc", bg='#667eea', fg='white',
                  command=self.filter_classes).pack(side=tk.LEFT, padx=5)
        tk.Button(filter_frame, text="🔄 Làm mới", bg='#e0e0e0',
                  command=self.refresh_classes).pack(side=tk.LEFT, padx=5)
        
        # Tree frame
        tree_frame = tk.Frame(tab, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        y_scroll = ttk.Scrollbar(tree_frame)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        x_scroll = ttk.Scrollbar(tree_frame, orient='horizontal')
        x_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.classes_tree = ttk.Treeview(
            tree_frame,
            columns=('id','code','name','teacher','credits','students','semester','year','status','created'),
            show='headings',
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set
        )
        y_scroll.config(command=self.classes_tree.yview)
        x_scroll.config(command=self.classes_tree.xview)
        
        headers = {
            'id': ('ID', 50),
            'code': ('Mã lớp', 100),
            'name': ('Tên lớp', 200),
            'teacher': ('Giảng viên', 150),
            'credits': ('Tín chỉ', 70),
            'students': ('SV', 70),
            'semester': ('Kỳ', 50),
            'year': ('Năm học', 100),
            'status': ('Trạng thái', 120),
            'created': ('Ngày tạo', 150)
        }
        for col, (text, width) in headers.items():
            self.classes_tree.heading(col, text=text)
            self.classes_tree.column(col, width=width)
        self.classes_tree.pack(fill=tk.BOTH, expand=True)
        
        # Action frame
        action_frame = tk.Frame(tab, bg='white')
        action_frame.pack(fill=tk.X, pady=10)
        
        tk.Button(action_frame, text="✓ Duyệt lớp", bg='#28a745', fg='white',
                  command=self.approve_selected, width=20).pack(side=tk.LEFT, padx=10)
        tk.Button(action_frame, text="✗ Từ chối lớp", bg='#dc3545', fg='white',
                  command=self.reject_selected, width=20).pack(side=tk.LEFT, padx=10)

    # =================== TAB 2: CẤP TÀI KHOẢN (SUB-TABS) ===================
    def create_account_creation_tab(self):
        tab = tk.Frame(self.main_notebook, bg='white')
        self.main_notebook.add(tab, text="➕ Cấp Tài Khoản")
        
        # Sub-notebook cho sinh viên và giảng viên
        sub_notebook = ttk.Notebook(tab)
        sub_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Sub-tab: Sinh viên
        self.create_account_creation_subtab(sub_notebook, 'student', '👥 Sinh Viên')
        
        # Sub-tab: Giảng viên
        self.create_account_creation_subtab(sub_notebook, 'teacher', '👩‍🏫 Giảng Viên')
    
    def create_account_creation_subtab(self, parent_notebook, account_type, tab_title):
        """Tạo sub-tab cho cấp tài khoản"""
        subtab = tk.Frame(parent_notebook, bg='white')
        parent_notebook.add(subtab, text=tab_title)
        
        # Thông tin hướng dẫn
        if account_type == 'student':
            headers = {'code':'MSSV','name':'Họ tên','gender':'Giới tính','dob':'Ngày sinh','major':'Ngành học','year':'Khóa học'}
            email_domain = 'student.edu.vn'
            columns = ['code','name','gender','dob','major','year']
        else:
            headers = {'code':'Mã GV','name':'Họ tên','gender':'Giới tính','dob':'Ngày sinh','faculty':'Khoa/Bộ môn'}
            email_domain = 'faculty.edu.vn'
            columns = ['code','name','gender','dob','faculty']
        
        info_frame = tk.LabelFrame(subtab, text="📖 Hướng dẫn", bg='white', font=('Arial', 11, 'bold'))
        info_frame.pack(fill=tk.X, padx=20, pady=10)
        info_text = f"""
• Chuẩn bị file Excel có các cột: {', '.join(headers.values())}
• Giới tính: male, female, other
• Ngày sinh: định dạng YYYY-MM-DD
• Hệ thống sẽ tự tạo: Username = {headers['code']}, Password = {headers['code']}, Email = {headers['code']}@{email_domain}
"""
        tk.Label(info_frame, text=info_text, bg='white', justify='left', font=('Arial', 10)).pack(padx=10, pady=10)
        
        # Buttons
        input_frame = tk.Frame(subtab, bg='white')
        input_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Button(input_frame, text="📂 Chọn file Excel", bg='#007bff', fg='white', font=('Arial', 12, 'bold'),
                  command=lambda:self.select_account_file(account_type), width=20).pack(side=tk.LEFT, padx=10)
        tk.Button(input_frame, text="➕ Thêm thủ công", bg='#17a2b8', fg='white', font=('Arial', 12, 'bold'),
                  command=lambda:self.add_account_manual(account_type), width=25).pack(side=tk.LEFT, padx=10)
        
        # Treeview
        tree_frame = tk.Frame(subtab, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        tk.Label(tree_frame, text=f"Danh sách {account_type} sẽ tạo:", bg='white', font=('Arial', 11, 'bold')).pack(anchor='w', pady=5)
        
        treeview = ttk.Treeview(tree_frame, columns=columns, show='headings', height=10)
        for col in columns:
            treeview.heading(col, text=headers[col])
            treeview.column(col, width=150)
        y_scroll = ttk.Scrollbar(tree_frame, orient='vertical', command=treeview.yview)
        treeview.configure(yscrollcommand=y_scroll.set)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        treeview.pack(fill=tk.BOTH, expand=True)
        
        # Action buttons
        action_frame = tk.Frame(subtab, bg='white')
        action_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Button(action_frame, text="✓ Tạo tài khoản", bg='#28a745', fg='white', font=('Arial', 12, 'bold'),
                  command=lambda:self.create_accounts(account_type), width=20).pack(side=tk.LEFT, padx=10)
        tk.Button(action_frame, text="📥 Xuất danh sách vừa tạo", bg='#ffc107', fg='black', font=('Arial', 11, 'bold'),
                  command=lambda:self.export_accounts_excel(account_type), width=25).pack(side=tk.LEFT, padx=10)
        tk.Button(action_frame, text="🗑️ Xóa danh sách", bg='#dc3545', fg='white', font=('Arial', 12, 'bold'),
                  command=lambda:self.clear_account_list(account_type), width=20).pack(side=tk.LEFT, padx=10)
        
        # Lưu reference
        setattr(self, f"{account_type}_creation_tree", treeview)

    # =================== TAB 3: QUẢN LÝ TÀI KHOẢN (SUB-TABS) ===================
    def create_account_management_tab(self):
        tab = tk.Frame(self.main_notebook, bg='white')
        self.main_notebook.add(tab, text="⚙️ Quản Lý Tài Khoản")
        
        # Sub-notebook
        sub_notebook = ttk.Notebook(tab)
        sub_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Sub-tab: Quản lý sinh viên
        self.create_account_management_subtab(sub_notebook, 'student', '👥 Sinh Viên')
        
        # Sub-tab: Quản lý giảng viên
        self.create_account_management_subtab(sub_notebook, 'teacher', '👩‍🏫 Giảng Viên')
    
    def create_account_management_subtab(self, parent_notebook, account_type, tab_title):
        """Tạo sub-tab quản lý tài khoản với tìm kiếm"""
        subtab = tk.Frame(parent_notebook, bg='white')
        parent_notebook.add(subtab, text=tab_title)
        
        # Search frame
        search_frame = tk.LabelFrame(subtab, text="🔍 Tìm kiếm", bg='white', font=('Arial', 11, 'bold'))
        search_frame.pack(fill=tk.X, padx=10, pady=10)
        
        search_row1 = tk.Frame(search_frame, bg='white')
        search_row1.pack(fill=tk.X, padx=10, pady=5)
        
        if account_type == 'student':
            tk.Label(search_row1, text="Mã SV:", bg='white').pack(side=tk.LEFT, padx=5)
            code_entry = tk.Entry(search_row1, width=15)
            code_entry.pack(side=tk.LEFT, padx=5)
            
            tk.Label(search_row1, text="Họ tên:", bg='white').pack(side=tk.LEFT, padx=5)
            name_entry = tk.Entry(search_row1, width=20)
            name_entry.pack(side=tk.LEFT, padx=5)
            
            tk.Label(search_row1, text="Ngành:", bg='white').pack(side=tk.LEFT, padx=5)
            major_entry = tk.Entry(search_row1, width=15)
            major_entry.pack(side=tk.LEFT, padx=5)
            
            tk.Label(search_row1, text="Khóa:", bg='white').pack(side=tk.LEFT, padx=5)
            year_entry = tk.Entry(search_row1, width=10)
            year_entry.pack(side=tk.LEFT, padx=5)
            
            search_entries = {
                'code': code_entry,
                'name': name_entry,
                'major': major_entry,
                'year': year_entry
            }
        else:  # teacher
            tk.Label(search_row1, text="Mã GV:", bg='white').pack(side=tk.LEFT, padx=5)
            code_entry = tk.Entry(search_row1, width=15)
            code_entry.pack(side=tk.LEFT, padx=5)
            
            tk.Label(search_row1, text="Họ tên:", bg='white').pack(side=tk.LEFT, padx=5)
            name_entry = tk.Entry(search_row1, width=20)
            name_entry.pack(side=tk.LEFT, padx=5)
            
            tk.Label(search_row1, text="Khoa/Bộ môn:", bg='white').pack(side=tk.LEFT, padx=5)
            dept_entry = tk.Entry(search_row1, width=20)
            dept_entry.pack(side=tk.LEFT, padx=5)
            
            search_entries = {
                'code': code_entry,
                'name': name_entry,
                'department': dept_entry
            }
        
        search_row2 = tk.Frame(search_frame, bg='white')
        search_row2.pack(fill=tk.X, padx=10, pady=5)
        
        tk.Button(search_row2, text="🔍 Tìm kiếm", bg='#007bff', fg='white',
                  command=lambda:self.search_accounts(account_type, search_entries), width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(search_row2, text="🔄 Hiển thị tất cả", bg='#6c757d', fg='white',
                  command=lambda:self.load_all_accounts(account_type), width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(search_row2, text="📤 Xuất Excel", bg='#28a745', fg='white',
                  command=lambda:self.export_all_from_database(account_type), width=15).pack(side=tk.LEFT, padx=5)
        
        # Treeview
        tree_frame = tk.Frame(subtab, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        if account_type == 'student':
            columns = ('id', 'code', 'name', 'gender', 'dob', 'email', 'major', 'year', 'face')
            headers = {
                'id': ('ID', 50),
                'code': ('Mã SV', 100),
                'name': ('Họ tên', 180),
                'gender': ('Giới tính', 80),
                'dob': ('Ngày sinh', 100),
                'email': ('Email', 200),
                'major': ('Ngành', 150),
                'year': ('Khóa', 70),
                'face': ('Ảnh', 60)
            }
        else:
            columns = ('id', 'code', 'name', 'gender', 'dob', 'email', 'department')
            headers = {
                'id': ('ID', 50),
                'code': ('Mã GV', 100),
                'name': ('Họ tên', 180),
                'gender': ('Giới tính', 80),
                'dob': ('Ngày sinh', 100),
                'email': ('Email', 200),
                'department': ('Khoa/Bộ môn', 200)
            }
        
        y_scroll = ttk.Scrollbar(tree_frame)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        x_scroll = ttk.Scrollbar(tree_frame, orient='horizontal')
        x_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        treeview = ttk.Treeview(
            tree_frame,
            columns=columns,
            show='headings',
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set
        )
        y_scroll.config(command=treeview.yview)
        x_scroll.config(command=treeview.xview)
        
        for col, (text, width) in headers.items():
            treeview.heading(col, text=text)
            treeview.column(col, width=width)
        treeview.pack(fill=tk.BOTH, expand=True)
        
        # Action buttons
        action_frame = tk.Frame(subtab, bg='white')
        action_frame.pack(fill=tk.X, padx=10, pady=10)
        
        tk.Button(action_frame, text="✏️ Cập nhật thông tin", bg='#ffc107', fg='black',
                  command=lambda:self.update_account(account_type, treeview), width=20).pack(side=tk.LEFT, padx=5)
        tk.Button(action_frame, text="🗑️ Xóa tài khoản", bg='#dc3545', fg='white',
                  command=lambda:self.delete_account(account_type, treeview), width=20).pack(side=tk.LEFT, padx=5)
        tk.Button(action_frame, text="🔄 Reset mật khẩu", bg='#17a2b8', fg='white',
                  command=lambda:self.reset_password(account_type, treeview), width=20).pack(side=tk.LEFT, padx=5)
        
        # Nút upload ảnh chỉ cho sinh viên
        if account_type == 'student':
            tk.Button(action_frame, text="📷 Upload ảnh khuôn mặt", bg='#6f42c1', fg='white',
                      command=lambda:self.upload_face_image(treeview), width=20).pack(side=tk.LEFT, padx=5)
            tk.Button(action_frame, text="👁️ Xem ảnh", bg='#20c997', fg='white',
                      command=lambda:self.view_face_image(treeview), width=15).pack(side=tk.LEFT, padx=5)
        
        # Lưu reference
        setattr(self, f"{account_type}_management_tree", treeview)
        setattr(self, f"{account_type}_search_entries", search_entries)
        
        # Load dữ liệu ban đầu
        self.load_all_accounts(account_type)

    # =================== TAB 4: QUẢN LÝ THỜI GIAN ĐĂNG KÝ ===================
    def create_registration_period_tab(self):
        tab = tk.Frame(self.main_notebook, bg='white')
        self.main_notebook.add(tab, text="⏰ Thời Gian Đăng Ký")
        
        frame = tk.Frame(tab, bg='white')
        frame.pack(padx=20, pady=20, anchor='w')
        
        tk.Label(frame, text="Học kỳ:", bg='white', font=('Arial', 11, 'bold')).grid(row=0, column=0, padx=5, pady=10, sticky='w')
        self.period_semester = ttk.Combobox(frame, values=['1', '2', '3'], width=10, state='readonly')
        self.period_semester.grid(row=0, column=1, padx=5, pady=10, sticky='w')
        self.period_semester.current(0)
        
        tk.Label(frame, text="Năm học:", bg='white', font=('Arial', 11, 'bold')).grid(row=0, column=2, padx=15, pady=10, sticky='w')
        self.period_year = tk.Entry(frame, width=12)
        self.period_year.grid(row=0, column=3, padx=5, pady=10, sticky='w')
        self.period_year.insert(0, "2024-2025")
        
        tk.Label(frame, text="Bắt đầu:", bg='white', font=('Arial', 11)).grid(row=1, column=0, padx=5, pady=5, sticky='w')
        self.start_date = DateEntry(
            frame, 
            width=12, 
            background='darkblue', 
            foreground='white', 
            borderwidth=2, 
            date_pattern='yyyy-mm-dd',
            showweeknumbers=False,
            showothermonthdays=True,
            selectbackground='#4472C4',
            selectforeground='white',
            state='normal',
            cursor='hand2'
        )
        self.start_date.grid(row=1, column=1, padx=5)
        self.start_hour = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(24)], width=3)
        self.start_hour.grid(row=1, column=2, padx=2)
        self.start_hour.set("08")
        self.start_minute = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(0,60,5)], width=3)
        self.start_minute.grid(row=1, column=3, padx=2)
        self.start_minute.set("00")
        
        tk.Label(frame, text="Kết thúc:", bg='white', font=('Arial', 11)).grid(row=2, column=0, padx=5, pady=5, sticky='w')
        self.end_date = DateEntry(
            frame, 
            width=12, 
            background='darkblue', 
            foreground='white', 
            borderwidth=2, 
            date_pattern='yyyy-mm-dd',
            showweeknumbers=False,
            showothermonthdays=True,
            selectbackground='#4472C4',
            selectforeground='white',
            state='normal',
            cursor='hand2'
        )
        self.end_date.grid(row=2, column=1, padx=5)
        self.end_hour = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(24)], width=3)
        self.end_hour.grid(row=2, column=2, padx=2)
        self.end_hour.set("23")
        self.end_minute = ttk.Combobox(frame, values=[f"{i:02d}" for i in range(0,60,5)], width=3)
        self.end_minute.grid(row=2, column=3, padx=2)
        self.end_minute.set("59")
        
        tk.Button(frame, text="💾 Lưu thời gian đăng ký", bg="#28a745", fg="white",
                  font=('Arial', 12, 'bold'), command=self.save_registration_period).grid(row=3, column=0, columnspan=4, pady=20)
    
    # =================== XỬ LÝ DUYỆT LỚP ===================
    def refresh_classes(self):
        for item in self.classes_tree.get_children():
            self.classes_tree.delete(item)
        
        try:
            classes = self.db.get_classes_for_approval()
            for c in classes:
                self.classes_tree.insert('', tk.END, values=(
                    c['class_id'], c['class_code'], c['class_name'], c['teacher_name'],
                    c['credits'], c['max_students'], c['semester'], c['academic_year'],
                    c['status'], c['created_at']
                ))
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không tải được lớp học:\n{e}")
    
    def filter_classes(self):
        semester = self.semester_combo.get()
        year = self.year_entry.get().strip()
        semester = None if semester == "Tất cả" else int(semester)
        year = None if not year else year
        
        for item in self.classes_tree.get_children():
            self.classes_tree.delete(item)
        
        try:
            classes = self.db.get_classes_for_approval(semester, year)
            for c in classes:
                self.classes_tree.insert('', tk.END, values=(
                    c['class_id'], c['class_code'], c['class_name'], c['teacher_name'],
                    c['credits'], c['max_students'], c['semester'], c['academic_year'],
                    c['status'], c['created_at']
                ))
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi lọc lớp:\n{e}")
    
    def approve_selected(self):
        selected = self.classes_tree.selection()
        if not selected:
            messagebox.showwarning("Chọn lớp", "Vui lòng chọn lớp cần duyệt")
            return
        class_id = self.classes_tree.item(selected[0])['values'][0]
        if self.db.approve_class(class_id):
            messagebox.showinfo("Thành công", "Đã duyệt lớp")
            self.refresh_classes()
        else:
            messagebox.showerror("Lỗi", "Không thể duyệt lớp")
    
    def reject_selected(self):
        selected = self.classes_tree.selection()
        if not selected:
            messagebox.showwarning("Chọn lớp", "Vui lòng chọn lớp cần từ chối")
            return
        class_id = self.classes_tree.item(selected[0])['values'][0]
        if self.db.reject_class(class_id):
            messagebox.showinfo("Thành công", "Đã từ chối lớp")
            self.refresh_classes()
        else:
            messagebox.showerror("Lỗi", "Không thể từ chối lớp")

    # =================== XỬ LÝ CẤP TÀI KHOẢN ===================
    def select_account_file(self, account_type):
        file_path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel files","*.xlsx *.xls")])
        if not file_path:
            return
        
        pending_list = getattr(self, f"pending_{account_type}s")
        treeview = getattr(self, f"{account_type}_creation_tree")
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            if account_type == 'student':
                keys = ['student_code','full_name','gender','date_of_birth','major','enrollment_year']
            else:
                keys = ['teacher_code','full_name','gender','date_of_birth','faculty']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                account = {k: (row[i] if i < len(row) and row[i] is not None else '') for i, k in enumerate(keys)}
                pending_list.append(account)
                treeview.insert('', tk.END, values=tuple(account.values()))
            messagebox.showinfo("Thành công", f"Đã tải {len(pending_list)} {account_type}!")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file: {e}")
    
    def add_account_manual(self, account_type):
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Thêm {account_type}")
        dialog.geometry("400x350")
        dialog.transient(self.root)
        dialog.grab_set()
        
        fields = {}
        if account_type == 'student':
            labels = [("MSSV:",'code'), ("Họ tên:",'full_name'), ("Giới tính:",'gender'),
                      ("Ngày sinh (YYYY-MM-DD):",'date_of_birth'), ("Ngành học:",'major'), ("Khóa học:",'year')]
        else:
            labels = [("Mã GV:",'code'), ("Họ tên:",'full_name'), ("Giới tính:",'gender'),
                      ("Ngày sinh (YYYY-MM-DD):",'date_of_birth'), ("Khoa/Bộ môn:",'faculty')]
        
        for i, (label_text, key) in enumerate(labels):
            tk.Label(dialog, text=label_text).grid(row=i, column=0, padx=10, pady=5, sticky='w')
            if key == 'gender':
                fields[key] = ttk.Combobox(dialog, values=['male','female','other'], width=25)
                fields[key].current(0)
            else:
                fields[key] = tk.Entry(dialog, width=28)
            fields[key].grid(row=i, column=1, padx=10, pady=5)
        
        def save_account():
            account = {key: fields[key].get().strip() for _, key in labels}
            if not account[list(account.keys())[0]] or not account['full_name']:
                messagebox.showerror("Lỗi", "ID và Họ tên không được để trống")
                return
            pending_list = getattr(self, f"pending_{account_type}s")
            treeview = getattr(self, f"{account_type}_creation_tree")
            pending_list.append(account)
            treeview.insert('', tk.END, values=tuple(account.values()))
            dialog.destroy()
        
        tk.Button(dialog, text="✓ Thêm", bg='#28a745', fg='white', command=save_account).grid(
            row=len(labels), column=0, columnspan=2, pady=20
        )
    
    def create_accounts(self, account_type):
        pending_list = getattr(self, f"pending_{account_type}s")
        if not pending_list:
            messagebox.showwarning("Cảnh báo", f"Chưa có {account_type} nào!")
            return
        if not messagebox.askyesno("Xác nhận", f"Tạo {len(pending_list)} tài khoản {account_type}?"):
            return
        
        func = {'student': self.db.create_students_bulk, 'teacher': self.db.create_teachers_bulk}[account_type]
        success_count, created = func(pending_list)
        setattr(self, f"created_{account_type}_accounts", created)
        messagebox.showinfo("Thành công", f"Đã tạo {success_count} tài khoản {account_type}!")
        
        # Refresh quản lý tài khoản
        self.load_all_accounts(account_type)
    
    def export_accounts_excel(self, account_type):
        created_accounts = getattr(self, f"created_{account_type}_accounts")
        if not created_accounts:
            messagebox.showwarning("Cảnh báo", f"Chưa có {account_type} nào để xuất!")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")])
        if not file_path:
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Accounts"
        keys = list(created_accounts[0].keys())
        for col, key in enumerate(keys, start=1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="00C0C0C0")
            cell.alignment = Alignment(horizontal='center')
        for r, account in enumerate(created_accounts, start=2):
            for c, key in enumerate(keys, start=1):
                ws.cell(row=r, column=c, value=account[key])
        wb.save(file_path)
        messagebox.showinfo("Thành công", f"Xuất Excel thành công: {file_path}")
    
    def clear_account_list(self, account_type):
        pending_list = getattr(self, f"pending_{account_type}s")
        treeview = getattr(self, f"{account_type}_creation_tree")
        if not pending_list:
            return
        if not messagebox.askyesno("Xác nhận", f"Xóa danh sách {account_type}?"):
            return
        pending_list.clear()
        for item in treeview.get_children():
            treeview.delete(item)

    # =================== XỬ LÝ QUẢN LÝ TÀI KHOẢN ===================
    def load_all_accounts(self, account_type):
        """Load tất cả tài khoản"""
        treeview = getattr(self, f"{account_type}_management_tree")
        
        # Xóa dữ liệu cũ
        for item in treeview.get_children():
            treeview.delete(item)
        
        try:
            if account_type == 'student':
                data = self.db.get_all_students()
                for item in data:
                    gender_map = {'male': 'Nam', 'female': 'Nữ', 'other': 'Khác'}
                    dob = item.get('date_of_birth')
                    dob_str = dob.strftime('%d/%m/%Y') if dob else ''
                    # Kiểm tra cả face_image (base64) và face_encoding_path (file)
                    has_face = "Có" if (item.get('face_image') or item.get('face_encoding_path')) else "Chưa"
                    
                    treeview.insert('', tk.END, values=(
                        item.get('student_id'),
                        item.get('student_code'),
                        item.get('full_name'),
                        gender_map.get(item.get('gender', ''), ''),
                        dob_str,
                        item.get('email'),
                        item.get('major', ''),
                        item.get('enrollment_year', ''),
                        has_face
                    ))
            else:  # teacher
                data = self.db.get_all_teachers()
                for item in data:
                    gender_map = {'male': 'Nam', 'female': 'Nữ', 'other': 'Khác'}
                    dob = item.get('date_of_birth')
                    dob_str = dob.strftime('%d/%m/%Y') if dob else ''
                    
                    treeview.insert('', tk.END, values=(
                        item.get('teacher_id'),
                        item.get('teacher_code'),
                        item.get('full_name'),
                        gender_map.get(item.get('gender', ''), ''),
                        dob_str,
                        item.get('email'),
                        item.get('department', '')
                    ))
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tải dữ liệu: {e}")
    
    def search_accounts(self, account_type, search_entries):
        """Tìm kiếm tài khoản"""
        treeview = getattr(self, f"{account_type}_management_tree")
        
        # Lấy giá trị tìm kiếm
        search_values = {key: entry.get().strip().lower() for key, entry in search_entries.items()}
        
        # Xóa dữ liệu cũ
        for item in treeview.get_children():
            treeview.delete(item)
        
        try:
            if account_type == 'student':
                data = self.db.get_all_students()
                for item in data:
                    # Kiểm tra điều kiện tìm kiếm
                    if search_values['code'] and search_values['code'] not in item.get('student_code', '').lower():
                        continue
                    if search_values['name'] and search_values['name'] not in item.get('full_name', '').lower():
                        continue
                    if search_values['major'] and search_values['major'] not in str(item.get('major', '')).lower():
                        continue
                    if search_values['year'] and search_values['year'] not in str(item.get('enrollment_year', '')):
                        continue
                    
                    gender_map = {'male': 'Nam', 'female': 'Nữ', 'other': 'Khác'}
                    dob = item.get('date_of_birth')
                    dob_str = dob.strftime('%d/%m/%Y') if dob else ''
                    # Kiểm tra cả face_image (base64) và face_encoding_path (file)
                    has_face = "Có" if (item.get('face_image') or item.get('face_encoding_path')) else "Chưa"
                    
                    treeview.insert('', tk.END, values=(
                        item.get('student_id'),
                        item.get('student_code'),
                        item.get('full_name'),
                        gender_map.get(item.get('gender', ''), ''),
                        dob_str,
                        item.get('email'),
                        item.get('major', ''),
                        item.get('enrollment_year', ''),
                        has_face
                    ))
            else:  # teacher
                data = self.db.get_all_teachers()
                for item in data:
                    # Kiểm tra điều kiện tìm kiếm
                    if search_values['code'] and search_values['code'] not in item.get('teacher_code', '').lower():
                        continue
                    if search_values['name'] and search_values['name'] not in item.get('full_name', '').lower():
                        continue
                    if search_values['department'] and search_values['department'] not in str(item.get('department', '')).lower():
                        continue
                    
                    gender_map = {'male': 'Nam', 'female': 'Nữ', 'other': 'Khác'}
                    dob = item.get('date_of_birth')
                    dob_str = dob.strftime('%d/%m/%Y') if dob else ''
                    
                    treeview.insert('', tk.END, values=(
                        item.get('teacher_id'),
                        item.get('teacher_code'),
                        item.get('full_name'),
                        gender_map.get(item.get('gender', ''), ''),
                        dob_str,
                        item.get('email'),
                        item.get('department', '')
                    ))
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi tìm kiếm: {e}")
    
    def update_account(self, account_type, treeview):
        """Cập nhật thông tin tài khoản với dialog mới"""
        from views.dialogs import UpdateStudentDialog, UpdateTeacherDialog
        
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Chọn tài khoản", f"Vui lòng chọn {account_type} cần cập nhật")
            return
        
        values = treeview.item(selected[0])['values']
        
        if account_type == 'student':
            # student_id, code, name, gender, dob, email, major, year, face
            data = {
                'id': values[0],
                'code': values[1],
                'name': values[2],
                'gender': values[3],
                'dob': values[4],
                'email': values[5],
                'phone': values[6] if len(values) > 6 else '',
                'major': values[7] if len(values) > 7 else '',
                'year': values[8] if len(values) > 8 else ''
            }
            UpdateStudentDialog(self.root, self.db, data, lambda: self.load_all_accounts('student'))
        else:
            # teacher_id, code, name, gender, dob, email, phone, department
            data = {
                'id': values[0],
                'code': values[1],
                'name': values[2],
                'gender': values[3],
                'dob': values[4],
                'email': values[5],
                'phone': values[6] if len(values) > 6 else '',
                'department': values[7] if len(values) > 7 else ''
            }
            UpdateTeacherDialog(self.root, self.db, data, lambda: self.load_all_accounts('teacher'))
    
    def _update_student_dialog(self, values, treeview):
        """Dialog cập nhật sinh viên"""
        student_id, code, name, gender, dob, email, major, year, _ = values
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Cập nhật sinh viên: {code}")
        dialog.geometry("450x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Form
        fields = {}
        labels = [
            ("Mã SV:", code, False),
            ("Họ tên:", name, True),
            ("Email:", email, True),
            ("Giới tính:", gender, True),
            ("Ngày sinh:", dob, True),
            ("Ngành học:", major, True),
            ("Khóa học:", year, True)
        ]
        
        for i, (label_text, default_value, editable) in enumerate(labels):
            tk.Label(dialog, text=label_text, font=('Arial', 10, 'bold')).grid(row=i, column=0, padx=10, pady=8, sticky='w')
            
            if label_text == "Giới tính:":
                gender_map = {'Nam': 'male', 'Nữ': 'female', 'Khác': 'other'}
                reverse_map = {v: k for k, v in gender_map.items()}
                fields['gender'] = ttk.Combobox(dialog, values=['Nam', 'Nữ', 'Khác'], width=30, state='readonly' if editable else 'disabled')
                fields['gender'].set(reverse_map.get(default_value, 'Nam'))
                fields['gender'].grid(row=i, column=1, padx=10, pady=8)
            else:
                entry = tk.Entry(dialog, width=33, state='normal' if editable else 'disabled')
                entry.insert(0, default_value)
                entry.grid(row=i, column=1, padx=10, pady=8)
                
                if label_text == "Họ tên:":
                    fields['full_name'] = entry
                elif label_text == "Email:":
                    fields['email'] = entry
                elif label_text == "Ngày sinh:":
                    fields['date_of_birth'] = entry
                elif label_text == "Ngành học:":
                    fields['major'] = entry
                elif label_text == "Khóa học:":
                    fields['enrollment_year'] = entry
        
        def save_update():
            from models.user import User
            from models.student import Student
            
            gender_map = {'Nam': 'male', 'Nữ': 'female', 'Khác': 'other'}
            
            # Lấy user_id từ student_id
            cursor = self.db.connection.cursor(dictionary=True)
            cursor.execute("SELECT user_id FROM students WHERE student_id = %s", (student_id,))
            result = cursor.fetchone()
            cursor.close()
            
            if not result:
                messagebox.showerror("Lỗi", "Không tìm thấy sinh viên")
                return
            
            user_id = result['user_id']
            
            # Cập nhật user
            user_model = User(self.db)
            user_model.update(
                user_id,
                full_name=fields['full_name'].get().strip(),
                email=fields['email'].get().strip(),
                gender=gender_map[fields['gender'].get()],
                date_of_birth=fields['date_of_birth'].get().strip()
            )
            
            # Cập nhật student
            student_model = Student(self.db)
            student_model.update(
                student_id,
                major=fields['major'].get().strip(),
                enrollment_year=int(fields['enrollment_year'].get().strip()) if fields['enrollment_year'].get().strip() else None
            )
            
            messagebox.showinfo("Thành công", "Đã cập nhật thông tin sinh viên")
            dialog.destroy()
            self.load_all_accounts('student')
        
        tk.Button(dialog, text="💾 Lưu", bg='#28a745', fg='white', font=('Arial', 12, 'bold'),
                  command=save_update, width=15).grid(row=len(labels), column=0, columnspan=2, pady=20)
    
    def _update_teacher_dialog(self, values, treeview):
        """Dialog cập nhật giảng viên"""
        teacher_id, code, name, gender, dob, email, department = values
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Cập nhật giảng viên: {code}")
        dialog.geometry("450x350")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Form
        fields = {}
        labels = [
            ("Mã GV:", code, False),
            ("Họ tên:", name, True),
            ("Email:", email, True),
            ("Giới tính:", gender, True),
            ("Ngày sinh:", dob, True),
            ("Khoa/Bộ môn:", department, True)
        ]
        
        for i, (label_text, default_value, editable) in enumerate(labels):
            tk.Label(dialog, text=label_text, font=('Arial', 10, 'bold')).grid(row=i, column=0, padx=10, pady=8, sticky='w')
            
            if label_text == "Giới tính:":
                gender_map = {'Nam': 'male', 'Nữ': 'female', 'Khác': 'other'}
                reverse_map = {v: k for k, v in gender_map.items()}
                fields['gender'] = ttk.Combobox(dialog, values=['Nam', 'Nữ', 'Khác'], width=30, state='readonly' if editable else 'disabled')
                fields['gender'].set(reverse_map.get(default_value, 'Nam'))
                fields['gender'].grid(row=i, column=1, padx=10, pady=8)
            else:
                entry = tk.Entry(dialog, width=33, state='normal' if editable else 'disabled')
                entry.insert(0, default_value)
                entry.grid(row=i, column=1, padx=10, pady=8)
                
                if label_text == "Họ tên:":
                    fields['full_name'] = entry
                elif label_text == "Email:":
                    fields['email'] = entry
                elif label_text == "Ngày sinh:":
                    fields['date_of_birth'] = entry
                elif label_text == "Khoa/Bộ môn:":
                    fields['department'] = entry
        
        def save_update():
            from models.user import User
            from models.teacher import Teacher
            
            gender_map = {'Nam': 'male', 'Nữ': 'female', 'Khác': 'other'}
            
            # Lấy user_id từ teacher_id
            cursor = self.db.connection.cursor(dictionary=True)
            cursor.execute("SELECT user_id FROM teachers WHERE teacher_id = %s", (teacher_id,))
            result = cursor.fetchone()
            cursor.close()
            
            if not result:
                messagebox.showerror("Lỗi", "Không tìm thấy giảng viên")
                return
            
            user_id = result['user_id']
            
            # Cập nhật user
            user_model = User(self.db)
            user_model.update(
                user_id,
                full_name=fields['full_name'].get().strip(),
                email=fields['email'].get().strip(),
                gender=gender_map[fields['gender'].get()],
                date_of_birth=fields['date_of_birth'].get().strip()
            )
            
            # Cập nhật teacher
            teacher_model = Teacher(self.db)
            teacher_model.update(
                teacher_id,
                department=fields['department'].get().strip()
            )
            
            messagebox.showinfo("Thành công", "Đã cập nhật thông tin giảng viên")
            dialog.destroy()
            self.load_all_accounts('teacher')
        
        tk.Button(dialog, text="💾 Lưu", bg='#28a745', fg='white', font=('Arial', 12, 'bold'),
                  command=save_update, width=15).grid(row=len(labels), column=0, columnspan=2, pady=20)
    
    def delete_account(self, account_type, treeview):
        """Xóa tài khoản"""
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Chọn tài khoản", f"Vui lòng chọn {account_type} cần xóa")
            return
        
        values = treeview.item(selected[0])['values']
        account_id = values[0]
        account_code = values[1]
        account_name = values[2]
        
        # Xác nhận xóa
        confirm_msg = f"⚠️ CẢNH BÁO: Xóa tài khoản sẽ xóa tất cả dữ liệu liên quan!\n\n"
        confirm_msg += f"Tài khoản: {account_name} ({account_code})\n\n"
        confirm_msg += "Bạn có chắc chắn muốn xóa?"
        
        if not messagebox.askyesno("Xác nhận xóa", confirm_msg):
            return
        
        try:
            from models.user import User
            
            # Lấy user_id
            cursor = self.db.connection.cursor(dictionary=True)
            if account_type == 'student':
                cursor.execute("SELECT user_id FROM students WHERE student_id = %s", (account_id,))
            else:
                cursor.execute("SELECT user_id FROM teachers WHERE teacher_id = %s", (account_id,))
            
            result = cursor.fetchone()
            cursor.close()
            
            if not result:
                messagebox.showerror("Lỗi", f"Không tìm thấy {account_type}")
                return
            
            user_id = result['user_id']
            
            # Xóa user (cascade sẽ xóa student/teacher)
            user_model = User(self.db)
            if user_model.delete(user_id):
                messagebox.showinfo("Thành công", f"Đã xóa {account_type}: {account_name}")
                self.load_all_accounts(account_type)
            else:
                messagebox.showerror("Lỗi", f"Không thể xóa {account_type}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi xóa tài khoản:\n{str(e)}")
    
    def reset_password(self, account_type, treeview):
        """Reset mật khẩu về mặc định"""
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Chọn tài khoản", f"Vui lòng chọn {account_type} cần reset mật khẩu")
            return
        
        values = treeview.item(selected[0])['values']
        account_id = values[0]
        account_code = values[1]
        account_name = values[2]
        
        confirm_msg = f"Reset mật khẩu về mặc định cho:\n\n"
        confirm_msg += f"Tài khoản: {account_name} ({account_code})\n"
        confirm_msg += f"Mật khẩu mới: {account_code}\n\n"
        confirm_msg += "Bạn có chắc chắn?"
        
        if not messagebox.askyesno("Xác nhận reset", confirm_msg):
            return
        
        try:
            from models.user import User
            
            # Lấy user_id
            cursor = self.db.connection.cursor(dictionary=True)
            if account_type == 'student':
                cursor.execute("SELECT user_id FROM students WHERE student_id = %s", (account_id,))
            else:
                cursor.execute("SELECT user_id FROM teachers WHERE teacher_id = %s", (account_id,))
            
            result = cursor.fetchone()
            cursor.close()
            
            if not result:
                messagebox.showerror("Lỗi", f"Không tìm thấy {account_type}")
                return
            
            user_id = result['user_id']
            
            # Reset password và đặt lại first_login = TRUE
            user_model = User(self.db)
            if user_model.reset_password(user_id, account_code):
                # Đặt lại first_login = TRUE để bắt buộc đổi mật khẩu
                cursor = self.db.connection.cursor()
                try:
                    cursor.execute("UPDATE users SET first_login = TRUE WHERE user_id = %s", (user_id,))
                    self.db.connection.commit()
                finally:
                    cursor.close()
                
                messagebox.showinfo("Thành công", 
                    f"Đã reset mật khẩu cho {account_name}\n\n"
                    f"Mật khẩu mới: {account_code}\n\n"
                    f"Người dùng sẽ được yêu cầu đổi mật khẩu khi đăng nhập lần đầu.")
            else:
                messagebox.showerror("Lỗi", "Không thể reset mật khẩu")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi reset mật khẩu:\n{str(e)}")
        messagebox.showinfo("Thông báo", f"Chức năng reset mật khẩu\n(Đang phát triển)")
    
    def export_all_from_database(self, account_type):
        """Xuất tất cả tài khoản ra Excel"""
        try:
            if account_type == 'student':
                data = self.db.get_all_students()
                title = "Danh Sách Sinh Viên"
                columns = [
                    ('STT', 'stt'),
                    ('Mã sinh viên', 'student_code'),
                    ('Họ và tên', 'full_name'),
                    ('Giới tính', 'gender'),
                    ('Ngày sinh', 'date_of_birth'),
                    ('Email', 'email'),
                    ('Ngành học', 'major'),
                    ('Khóa học', 'enrollment_year'),
                    ('Có ảnh khuôn mặt', 'has_face')
                ]
            else:
                data = self.db.get_all_teachers()
                title = "Danh Sách Giảng Viên"
                columns = [
                    ('STT', 'stt'),
                    ('Mã giảng viên', 'teacher_code'),
                    ('Họ và tên', 'full_name'),
                    ('Giới tính', 'gender'),
                    ('Ngày sinh', 'date_of_birth'),
                    ('Email', 'email'),
                    ('Khoa/Bộ môn', 'department')
                ]
            
            if not data:
                messagebox.showwarning("Cảnh báo", f"Không có {account_type} nào trong database!")
                return
            
            default_filename = f"DanhSach_{account_type.capitalize()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=default_filename
            )
            
            if not file_path:
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_idx, (header_text, _) in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            for row_idx, item in enumerate(data, start=2):
                for col_idx, (_, key) in enumerate(columns, start=1):
                    if key == 'stt':
                        value = row_idx - 1
                    elif key == 'has_face':
                        value = "Có" if item.get('face_encoding_path') else "Chưa"
                    elif key == 'date_of_birth':
                        dob = item.get(key)
                        value = dob.strftime('%d/%m/%Y') if dob else ''
                    elif key == 'gender':
                        gender_map = {'male': 'Nam', 'female': 'Nữ', 'other': 'Khác'}
                        value = gender_map.get(item.get(key, ''), item.get(key, ''))
                    else:
                        value = item.get(key, '')
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            for col_idx, (header_text, _) in enumerate(columns, start=1):
                max_length = len(header_text)
                for row_idx in range(2, len(data) + 2):
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value)
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
            
            summary_row = len(data) + 3
            ws.cell(row=summary_row, column=1, value=f"Tổng số: {len(data)} {account_type}")
            ws.cell(row=summary_row, column=1).font = Font(bold=True, size=11)
            ws.cell(row=summary_row + 1, column=1, value=f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            ws.cell(row=summary_row + 1, column=1).font = Font(italic=True, size=10)
            
            wb.save(file_path)
            messagebox.showinfo("Thành công", f"Đã xuất {len(data)} {account_type} ra file:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất Excel:\n{str(e)}")
    
    # =================== QUẢN LÝ THỜI GIAN ĐĂNG KÝ ===================
    def save_registration_period(self):
        try:
            start_dt = datetime.combine(
                self.start_date.get_date(),
                time(int(self.start_hour.get()), int(self.start_minute.get()))
            )
            end_dt = datetime.combine(
                self.end_date.get_date(),
                time(int(self.end_hour.get()), int(self.end_minute.get()))
            )
            
            semester = int(self.period_semester.get())
            year = self.period_year.get().strip()
            
            if start_dt >= end_dt:
                messagebox.showerror("Lỗi", "Thời gian bắt đầu phải trước kết thúc")
                return
            
            if self.db.save_registration_period(start_dt, end_dt, semester, year):
                messagebox.showinfo("Thành công", "Đã lưu thời gian đăng ký")
            else:
                messagebox.showerror("Lỗi", "Không thể lưu thời gian")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Dữ liệu không hợp lệ:\n{e}")
    
    def load_registration_period(self):
        data = self.db.get_latest_registration_window()
        if data:
            self.period_semester.set(str(data['semester']))
            self.period_year.delete(0, tk.END)
            self.period_year.insert(0, data['academic_year'])
            self.start_date.set_date(data['start_datetime'])
            self.end_date.set_date(data['end_datetime'])
            self.start_hour.set(f"{data['start_datetime'].hour:02d}")
            self.start_minute.set(f"{data['start_datetime'].minute:02d}")
            self.end_hour.set(f"{data['end_datetime'].hour:02d}")
            self.end_minute.set(f"{data['end_datetime'].minute:02d}")
    
    # =================== LOGOUT ===================
    def logout(self):
        if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn đăng xuất?"):
            self.logout_callback()

    
    # =================== XỬ LÝ ẢNH KHUÔN MẶT ===================
    def upload_face_image(self, treeview):
        """Upload ảnh khuôn mặt cho sinh viên - Chọn từ file hoặc webcam"""
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Chọn sinh viên", "Vui lòng chọn sinh viên cần upload ảnh")
            return
        
        values = treeview.item(selected[0])['values']
        student_id = values[0]
        student_code = values[1]
        student_name = values[2]
        
        # Dialog chọn phương thức
        choice_dialog = tk.Toplevel(self.root)
        choice_dialog.title("Chọn phương thức upload")
        choice_dialog.geometry("400x250")
        choice_dialog.transient(self.root)
        choice_dialog.grab_set()
        choice_dialog.resizable(False, False)
        
        # Center dialog
        choice_dialog.update_idletasks()
        x = (choice_dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (choice_dialog.winfo_screenheight() // 2) - (250 // 2)
        choice_dialog.geometry(f"400x250+{x}+{y}")
        
        tk.Label(
            choice_dialog,
            text=f"Upload ảnh cho: {student_name}",
            font=('Arial', 12, 'bold'),
            bg='white'
        ).pack(pady=20)
        
        tk.Label(
            choice_dialog,
            text="Chọn phương thức upload ảnh:",
            font=('Arial', 10),
            bg='white'
        ).pack(pady=10)
        
        def from_file():
            choice_dialog.destroy()
            # Chọn file ảnh
            file_path = filedialog.askopenfilename(
                title="Chọn ảnh khuôn mặt",
                filetypes=[
                    ("Image files", "*.jpg *.jpeg *.png *.bmp"),
                    ("All files", "*.*")
                ]
            )
            
            if not file_path:
                return
            
            # Validate và upload
            from utils.image_handler import FaceImageDB
            
            face_db = FaceImageDB(self.db)
            success, message = face_db.save_face_image(student_id, file_path, compress=True)
            
            if success:
                messagebox.showinfo("Thành công", f"Đã upload ảnh cho {student_name}\n{message}")
                self.load_all_accounts('student')
            else:
                messagebox.showerror("Lỗi", f"Không thể upload ảnh:\n{message}")
        
        def from_webcam():
            choice_dialog.destroy()
            from utils.webcam_capture import WebcamCapture
            import cv2
            
            def on_capture(frame):
                """Callback khi chụp ảnh xong"""
                # Lưu frame vào file tạm
                temp_file = WebcamCapture.save_frame_to_temp(frame)
                
                # Upload vào database
                from utils.image_handler import FaceImageDB
                face_db = FaceImageDB(self.db)
                success, message = face_db.save_face_image(student_id, temp_file, compress=True)
                
                if success:
                    messagebox.showinfo("Thành công", f"Đã upload ảnh cho {student_name}\n{message}")
                    self.load_all_accounts('student')
                else:
                    messagebox.showerror("Lỗi", f"Không thể upload ảnh:\n{message}")
                
                # Xóa file tạm
                import os
                try:
                    os.remove(temp_file)
                except:
                    pass
            
            # Mở webcam capture
            webcam = WebcamCapture(self.root, on_capture, f"Chụp ảnh cho {student_name}")
            webcam.open_camera()
        
        # Buttons
        btn_frame = tk.Frame(choice_dialog, bg='white')
        btn_frame.pack(pady=20)
        
        tk.Button(
            btn_frame,
            text="📂 Chọn từ máy tính",
            font=('Arial', 11, 'bold'),
            bg='#007bff',
            fg='white',
            command=from_file,
            width=18
        ).pack(side=tk.LEFT, padx=10)
        
        tk.Button(
            btn_frame,
            text="📸 Chụp từ Webcam",
            font=('Arial', 11, 'bold'),
            bg='#6f42c1',
            fg='white',
            command=from_webcam,
            width=18
        ).pack(side=tk.LEFT, padx=10)
    
    def view_face_image(self, treeview):
        """Xem ảnh khuôn mặt của sinh viên"""
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Chọn sinh viên", "Vui lòng chọn sinh viên cần xem ảnh")
            return
        
        values = treeview.item(selected[0])['values']
        student_id = values[0]
        student_code = values[1]
        student_name = values[2]
        has_face = values[8] if len(values) > 8 else "Chưa"
        
        if has_face != "Có":
            messagebox.showinfo("Thông báo", f"{student_name} chưa có ảnh khuôn mặt")
            return
        
        # Lấy ảnh từ database
        from utils.image_handler import FaceImageDB
        import cv2
        
        face_db = FaceImageDB(self.db)
        image = face_db.get_face_image(student_id, as_array=True)
        
        if image is None:
            messagebox.showerror("Lỗi", "Không thể tải ảnh từ database")
            return
        
        # Hiển thị ảnh trong dialog
        self._show_image_dialog(image, student_name, student_code, student_id)
    
    def _show_image_dialog(self, image, student_name, student_code, student_id):
        """Hiển thị dialog với ảnh khuôn mặt"""
        import cv2
        from PIL import Image, ImageTk
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Ảnh khuôn mặt: {student_name} ({student_code})")
        dialog.geometry("600x700")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Info frame
        info_frame = tk.Frame(dialog, bg='white')
        info_frame.pack(fill=tk.X, padx=10, pady=10)
        
        tk.Label(info_frame, text=f"Sinh viên: {student_name}", bg='white', 
                 font=('Arial', 12, 'bold')).pack(anchor='w')
        tk.Label(info_frame, text=f"Mã SV: {student_code}", bg='white', 
                 font=('Arial', 11)).pack(anchor='w')
        
        # Image frame
        image_frame = tk.Frame(dialog, bg='white')
        image_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Resize ảnh để hiển thị
        height, width = image.shape[:2]
        max_size = 500
        if width > max_size or height > max_size:
            scale = min(max_size / width, max_size / height)
            new_width = int(width * scale)
            new_height = int(height * scale)
            display_image = cv2.resize(image, (new_width, new_height))
        else:
            display_image = image
        
        # Chuyển BGR sang RGB
        display_image = cv2.cvtColor(display_image, cv2.COLOR_BGR2RGB)
        
        # Chuyển sang PIL Image
        pil_image = Image.fromarray(display_image)
        photo = ImageTk.PhotoImage(pil_image)
        
        # Label hiển thị ảnh
        image_label = tk.Label(image_frame, image=photo, bg='white')
        image_label.image = photo  # Giữ reference
        image_label.pack()
        
        # Info text
        from utils.image_handler import ImageHandler
        base64_str = FaceImageDB(self.db).get_face_image(student_id, as_array=False)
        if base64_str:
            info = ImageHandler.get_image_info(base64_str)
            info_text = f"Kích thước: {info.get('width')}x{info.get('height')} | "
            info_text += f"Dung lượng: {info.get('size_kb')} KB"
            tk.Label(dialog, text=info_text, bg='white', font=('Arial', 9), 
                     fg='gray').pack(pady=5)
        
        # Action buttons
        action_frame = tk.Frame(dialog, bg='white')
        action_frame.pack(fill=tk.X, padx=10, pady=10)
        
        def delete_image():
            if messagebox.askyesno("Xác nhận", f"Xóa ảnh khuôn mặt của {student_name}?"):
                from utils.image_handler import FaceImageDB
                face_db = FaceImageDB(self.db)
                success, msg = face_db.delete_face_image(student_id)
                if success:
                    messagebox.showinfo("Thành công", msg)
                    dialog.destroy()
                    self.load_all_accounts('student')
                else:
                    messagebox.showerror("Lỗi", msg)
        
        def export_image():
            file_path = filedialog.asksaveasfilename(
                defaultextension=".jpg",
                filetypes=[("JPEG files", "*.jpg"), ("PNG files", "*.png")],
                initialfile=f"{student_code}_face.jpg"
            )
            if file_path:
                cv2.imwrite(file_path, image)
                messagebox.showinfo("Thành công", f"Đã xuất ảnh: {file_path}")
        
        tk.Button(action_frame, text="🗑️ Xóa ảnh", bg='#dc3545', fg='white',
                  command=delete_image, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(action_frame, text="💾 Xuất ảnh", bg='#28a745', fg='white',
                  command=export_image, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(action_frame, text="✖️ Đóng", bg='#6c757d', fg='white',
                  command=dialog.destroy, width=15).pack(side=tk.LEFT, padx=5)
    
    # =================== TAB 5: TRAIN MODEL AI ===================
    def create_train_model_tab(self):
        """Tab train model AI"""
        tab = tk.Frame(self.main_notebook, bg='white')
        self.main_notebook.add(tab, text='🤖 Train Model AI')

        # Info frame
        info_frame = tk.LabelFrame(tab, text="📖 Hướng dẫn", bg='white', font=('Arial', 11, 'bold'))
        info_frame.pack(fill=tk.X, padx=20, pady=10)

        info_text = """
• Train model AI để hệ thống có thể nhận diện khuôn mặt sinh viên
• Yêu cầu: Mỗi sinh viên cần có ít nhất 1 ảnh khuôn mặt
• Thời gian train: 1-5 phút tùy số lượng sinh viên
• Nên train lại khi:
  - Có sinh viên mới upload ảnh
  - Nhận được yêu cầu từ giảng viên
  - Hệ thống nhận diện không chính xác
        """
        tk.Label(info_frame, text=info_text, bg='white', justify='left', 
                 font=('Arial', 10)).pack(padx=10, pady=10)

        # Stats frame
        stats_frame = tk.LabelFrame(tab, text="📊 Thống kê", bg='white', font=('Arial', 11, 'bold'))
        stats_frame.pack(fill=tk.X, padx=20, pady=10)

        self.train_stats_label = tk.Label(
            stats_frame,
            text="Đang tải thống kê...",
            font=('Arial', 10),
            bg='white',
            justify='left'
        )
        self.train_stats_label.pack(padx=10, pady=10)

        # Train button frame
        train_frame = tk.Frame(tab, bg='white')
        train_frame.pack(expand=True)

        self.train_model_btn = tk.Button(
            train_frame,
            text="🚀 Train Model AI",
            font=('Arial', 14, 'bold'),
            bg='#667eea',
            fg='white',
            cursor='hand2',
            command=self.train_ai_model,
            width=25,
            height=2
        )
        self.train_model_btn.pack(pady=20)

        # Status label
        self.train_model_status = tk.Label(
            train_frame,
            text="",
            font=('Arial', 11),
            bg='white',
            fg='#666'
        )
        self.train_model_status.pack(pady=10)

        # Load stats
        self.load_train_stats()

    def load_train_stats(self):
        """Load thống kê để hiển thị"""
        try:
            cursor = self.db.connection.cursor(dictionary=True)
            
            # Đếm sinh viên có ảnh
            cursor.execute("""
                SELECT COUNT(*) as total FROM students 
                WHERE face_image IS NOT NULL AND face_image != ''
            """)
            students_with_image = cursor.fetchone()['total']
            
            # Đếm tổng sinh viên
            cursor.execute("SELECT COUNT(*) as total FROM students")
            total_students = cursor.fetchone()['total']
            
            cursor.close()

            stats_text = f"""
Tổng số sinh viên: {total_students}
Sinh viên có ảnh: {students_with_image}
Sinh viên chưa có ảnh: {total_students - students_with_image}
Tỷ lệ: {(students_with_image/total_students*100) if total_students > 0 else 0:.1f}%
            """
            self.train_stats_label.config(text=stats_text)

        except Exception as e:
            self.train_stats_label.config(text=f"Lỗi tải thống kê: {e}")

    def train_ai_model(self):
        """Train AI model"""
        if not messagebox.askyesno("Xác nhận", 
            "Train model AI?\n\n"
            "Quá trình này có thể mất vài phút.\n"
            "Bạn có muốn tiếp tục?"):
            return

        try:
            self.train_model_btn.config(state=tk.DISABLED, text="⏳ Đang train...")
            self.train_model_status.config(text="Vui lòng đợi...", fg='#666')
            self.root.update()

            from services.face_recognition_service import face_service
            result = face_service.train_model(self.db)

            if result.get('success'):
                self.train_model_status.config(
                    text=f"✓ Thành công! Đã train {result.get('total_students',0)} sinh viên với {result.get('total_images',0)} ảnh",
                    fg='#48bb78'
                )
                messagebox.showinfo("Thành công", 
                    f"Model AI đã được train thành công!\n\n"
                    f"Sinh viên: {result.get('total_students',0)}\n"
                    f"Tổng ảnh: {result.get('total_images',0)}\n\n"
                    f"Hệ thống đã sẵn sàng cho điểm danh tự động!")
                
                # Reload stats
                self.load_train_stats()
            else:
                self.train_model_status.config(
                    text=f"✗ Lỗi: {result.get('error', 'Unknown')}",
                    fg='#f56565'
                )
                messagebox.showerror("Lỗi", result.get('error', 'Không thể train model'))

        except Exception as e:
            self.train_model_status.config(text=f"✗ Lỗi: {str(e)}", fg='#f56565')
            messagebox.showerror("Lỗi", f"Không thể train model:\n{str(e)}")
        finally:
            self.train_model_btn.config(state=tk.NORMAL, text="🚀 Train Model AI")

    # =================== TAB 6: YÊU CẦU HỖ TRỢ ===================
    def create_support_requests_tab(self):
        """Tab xem yêu cầu hỗ trợ từ giảng viên"""
        tab = tk.Frame(self.main_notebook, bg='white')
        self.main_notebook.add(tab, text='📨 Yêu cầu hỗ trợ')

        # Toolbar
        toolbar = tk.Frame(tab, bg='white')
        toolbar.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(
            toolbar,
            text="Danh sách yêu cầu từ giảng viên:",
            font=('Arial', 11, 'bold'),
            bg='white'
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            toolbar,
            text="🔄 Làm mới",
            bg='#17a2b8',
            fg='white',
            font=('Arial', 10, 'bold'),
            command=self.refresh_support_requests,
            width=15
        ).pack(side=tk.RIGHT, padx=5)

        # Treeview
        tree_frame = tk.Frame(tab, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        y_scroll = ttk.Scrollbar(tree_frame)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.support_requests_tree = ttk.Treeview(
            tree_frame,
            columns=('id', 'teacher', 'type', 'content', 'status', 'date'),
            show='headings',
            yscrollcommand=y_scroll.set
        )
        y_scroll.config(command=self.support_requests_tree.yview)

        headers = {
            'id': ('ID', 50),
            'teacher': ('Giảng viên', 150),
            'type': ('Loại yêu cầu', 150),
            'content': ('Nội dung', 300),
            'status': ('Trạng thái', 120),
            'date': ('Ngày gửi', 150)
        }

        for col, (text, width) in headers.items():
            self.support_requests_tree.heading(col, text=text)
            self.support_requests_tree.column(col, width=width)

        self.support_requests_tree.pack(fill=tk.BOTH, expand=True)

        # Action buttons
        action_frame = tk.Frame(tab, bg='white')
        action_frame.pack(pady=10)

        tk.Button(
            action_frame,
            text="👁️ Xem chi tiết",
            bg='#17a2b8',
            fg='white',
            font=('Arial', 11, 'bold'),
            command=self.view_support_request_detail,
            width=18
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            action_frame,
            text="✅ Đánh dấu hoàn thành",
            bg='#28a745',
            fg='white',
            font=('Arial', 11, 'bold'),
            command=self.mark_request_completed,
            width=20
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            action_frame,
            text="🗑️ Xóa yêu cầu",
            bg='#dc3545',
            fg='white',
            font=('Arial', 11, 'bold'),
            command=self.delete_support_request,
            width=18
        ).pack(side=tk.LEFT, padx=5)

        # Load data
        self.refresh_support_requests()

    def refresh_support_requests(self):
        """Làm mới danh sách yêu cầu"""
        try:
            for item in self.support_requests_tree.get_children():
                self.support_requests_tree.delete(item)

            cursor = self.db.connection.cursor(dictionary=True)
            cursor.execute("""
                SELECT sr.*, u.full_name as teacher_name
                FROM support_requests sr
                JOIN teachers t ON sr.teacher_id = t.teacher_id
                JOIN users u ON t.user_id = u.user_id
                ORDER BY sr.created_at DESC
            """)
            requests = cursor.fetchall()
            cursor.close()

            type_map = {
                'update_ai': '🤖 Cập nhật AI',
                'recognition_error': '❌ Lỗi nhận diện',
                'new_student': '➕ Sinh viên mới',
                'other': '❓ Khác'
            }

            status_map = {
                'pending': '⏳ Chờ xử lý',
                'processing': '🔄 Đang xử lý',
                'completed': '✅ Hoàn thành'
            }

            for req in requests:
                self.support_requests_tree.insert('', tk.END, values=(
                    req['request_id'],
                    req['teacher_name'],
                    type_map.get(req['request_type'], req['request_type']),
                    req['content'][:50] + '...' if len(req['content']) > 50 else req['content'],
                    status_map.get(req['status'], req['status']),
                    req['created_at'].strftime('%d/%m/%Y %H:%M')
                ))

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tải yêu cầu:\n{str(e)}")

    def view_support_request_detail(self):
        """Xem chi tiết yêu cầu"""
        selected = self.support_requests_tree.selection()
        if not selected:
            messagebox.showwarning("Chọn yêu cầu", "Vui lòng chọn yêu cầu cần xem!")
            return

        request_id = self.support_requests_tree.item(selected[0])['values'][0]

        try:
            cursor = self.db.connection.cursor(dictionary=True)
            cursor.execute("""
                SELECT sr.*, u.full_name as teacher_name, u.email as teacher_email
                FROM support_requests sr
                JOIN teachers t ON sr.teacher_id = t.teacher_id
                JOIN users u ON t.user_id = u.user_id
                WHERE sr.request_id = %s
            """, (request_id,))
            req = cursor.fetchone()
            cursor.close()

            if not req:
                messagebox.showerror("Lỗi", "Không tìm thấy yêu cầu!")
                return

            # Dialog hiển thị
            dialog = tk.Toplevel(self.root)
            dialog.title(f"Chi tiết yêu cầu #{request_id}")
            dialog.geometry("600x500")
            dialog.transient(self.root)
            dialog.grab_set()

            # Center
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - 300
            y = (dialog.winfo_screenheight() // 2) - 250
            dialog.geometry(f'600x500+{x}+{y}')

            # Content
            content = tk.Frame(dialog, bg='white', padx=20, pady=20)
            content.pack(fill=tk.BOTH, expand=True)

            tk.Label(content, text=f"Yêu cầu #{request_id}", 
                     font=('Arial', 14, 'bold'), bg='white').pack(anchor='w', pady=5)

            info = [
                ("Giảng viên:", req['teacher_name']),
                ("Email:", req['teacher_email']),
                ("Loại yêu cầu:", req['request_type']),
                ("Trạng thái:", req['status']),
                ("Ngày gửi:", req['created_at'].strftime('%d/%m/%Y %H:%M'))
            ]

            for label, value in info:
                row = tk.Frame(content, bg='white')
                row.pack(fill=tk.X, pady=3)
                tk.Label(row, text=label, font=('Arial', 10, 'bold'), 
                         bg='white', width=15, anchor='w').pack(side=tk.LEFT)
                tk.Label(row, text=value, font=('Arial', 10), 
                         bg='white', anchor='w').pack(side=tk.LEFT)

            tk.Label(content, text="Nội dung:", font=('Arial', 10, 'bold'), 
                     bg='white').pack(anchor='w', pady=(10,5))

            text_widget = tk.Text(content, font=('Arial', 10), height=10, 
                                  relief=tk.SOLID, borderwidth=1)
            text_widget.insert('1.0', req['content'])
            text_widget.config(state=tk.DISABLED)
            text_widget.pack(fill=tk.BOTH, expand=True)

            tk.Button(dialog, text="✖️ Đóng", bg='#6c757d', fg='white',
                      command=dialog.destroy, width=15).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xem chi tiết:\n{str(e)}")

    def mark_request_completed(self):
        """Đánh dấu yêu cầu hoàn thành"""
        selected = self.support_requests_tree.selection()
        if not selected:
            messagebox.showwarning("Chọn yêu cầu", "Vui lòng chọn yêu cầu!")
            return

        request_id = self.support_requests_tree.item(selected[0])['values'][0]

        if not messagebox.askyesno("Xác nhận", "Đánh dấu yêu cầu này đã hoàn thành?"):
            return

        try:
            cursor = self.db.connection.cursor()
            cursor.execute("""
                UPDATE support_requests 
                SET status = 'completed'
                WHERE request_id = %s
            """, (request_id,))
            self.db.connection.commit()
            cursor.close()

            messagebox.showinfo("Thành công", "Đã đánh dấu hoàn thành!")
            self.refresh_support_requests()

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể cập nhật:\n{str(e)}")

    def delete_support_request(self):
        """Xóa yêu cầu"""
        selected = self.support_requests_tree.selection()
        if not selected:
            messagebox.showwarning("Chọn yêu cầu", "Vui lòng chọn yêu cầu cần xóa!")
            return

        request_id = self.support_requests_tree.item(selected[0])['values'][0]

        if not messagebox.askyesno("Xác nhận", "Xóa yêu cầu này?"):
            return

        try:
            cursor = self.db.connection.cursor()
            cursor.execute("DELETE FROM support_requests WHERE request_id = %s", (request_id,))
            self.db.connection.commit()
            cursor.close()

            messagebox.showinfo("Thành công", "Đã xóa yêu cầu!")
            self.refresh_support_requests()

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xóa:\n{str(e)}")

    # =================== XỬ LÝ ĐÓNG CỬA SỔ ===================
    def on_closing(self):
        """Xử lý khi đóng cửa sổ"""
        if messagebox.askyesno("Xác nhận thoát", "Bạn có chắc muốn thoát?"):
            self.root.destroy()
            import sys
            sys.exit(0)
